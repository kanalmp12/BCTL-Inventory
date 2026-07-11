import openpyxl
import psycopg2
import os

file_path = "/Users/koratach/Desktop/BCTL-Inventory/BCTL-Inventory/Tool Crib Inventory System DB.xlsx"
connection_string = "postgresql://postgres:BZslqE9FxoSWJKOZ@db.ctypyzwlmzmawhnlsylf.supabase.co:5432/postgres"

def get_category(tool_id):
    if not tool_id:
        return 'Decoration & Props'
    prefix = tool_id[0].upper()

    # Cabinet → Category mapping (based on actual item types per cabinet)
    # A (ตู้1), H (ตู้8), K (ตู้11), O (บนตู้), N (ตู้14) → Decoration & Props
    if prefix in ['A', 'H', 'K', 'O', 'N']:
        return 'Decoration & Props'
    # B (ตู้2), E (ตู้5) → Stationery & Craft
    elif prefix in ['B', 'E']:
        return 'Stationery & Craft'
    # C (ตู้3), D (ตู้4), J (ตู้10) → Clothing & Fabric
    elif prefix in ['C', 'D', 'J']:
        return 'Clothing & Fabric'
    # G (ตู้7) → Catering & Cleaning
    elif prefix in ['G']:
        return 'Catering & Cleaning'
    # F (ตู้6), L (ตู้12), M (ตู้13) → AV & Lighting
    elif prefix in ['F', 'L', 'M']:
        return 'AV & Lighting'
    # I (ตู้9), P (ด้านใน) → Storage & Infrastructure
    elif prefix in ['I', 'P']:
        return 'Storage & Infrastructure'
    else:
        return 'Decoration & Props'

def seed():
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['Inventory']
    
    conn = psycopg2.connect(connection_string)
    cur = conn.cursor()
    
    inserted_count = 0
    
    try:
        print("Connected to Supabase. Seeding items...")
        for row_idx in range(2, sheet.max_row + 1):
            tool_id = sheet.cell(row=row_idx, column=1).value
            name = sheet.cell(row=row_idx, column=2).value
            total_qty_val = sheet.cell(row=row_idx, column=3).value
            avail_qty_val = sheet.cell(row=row_idx, column=4).value
            unit = sheet.cell(row=row_idx, column=5).value
            location = sheet.cell(row=row_idx, column=6).value
            image_url = sheet.cell(row=row_idx, column=7).value
            
            if not tool_id or not name:
                continue
                
            # Process quantities
            is_many = False
            if str(total_qty_val).strip() == 'จำนวนมาก' or str(avail_qty_val).strip() == 'จำนวนมาก':
                is_many = True
                total_qty = 1
                available_qty = 1
            else:
                try:
                    total_qty = int(float(total_qty_val))
                    available_qty = int(float(avail_qty_val))
                except (ValueError, TypeError):
                    total_qty = 1
                    available_qty = 1
            
            category = get_category(tool_id)
            # Append unit in Thai name if applicable, or keep name clean
            name_th = name
            name_en = name
            
            # Clean image URL if empty/None
            if image_url and not str(image_url).startswith('http'):
                image_url = None
                
            if not location:
                location = "ตู้เก็บของ"
                
            cur.execute("""
                INSERT INTO public.inventory_items (
                    id, name_th, name_en, category, location, total_qty, available_qty, is_many, image_url, is_active, created_at, updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, true, now(), now()
                )
                ON CONFLICT (id) DO UPDATE SET
                    name_th = EXCLUDED.name_th,
                    name_en = EXCLUDED.name_en,
                    category = EXCLUDED.category,
                    location = EXCLUDED.location,
                    total_qty = EXCLUDED.total_qty,
                    available_qty = EXCLUDED.available_qty,
                    is_many = EXCLUDED.is_many,
                    image_url = COALESCE(EXCLUDED.image_url, public.inventory_items.image_url),
                    updated_at = now()
            """, (tool_id, name_th, name_en, category, location, total_qty, available_qty, is_many, image_url))
            
            inserted_count += 1
            
        conn.commit()
        print(f"Seeding completed successfully! Total items seeded: {inserted_count}")
        
    except Exception as e:
        conn.rollback()
        print("Seeding failed:", e)
    finally:
        cur.close()
        conn.close()

if __name__ == "__main__":
    seed()
